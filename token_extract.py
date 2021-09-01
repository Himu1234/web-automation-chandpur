import os
import re
import selenium
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

excel_file = 'token_generation_automation.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(), excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[2]]
max_consumers = ws1.max_row
print(max_consumers)

browser = webdriver.Chrome(executable_path = os.path.join(os.getcwd(), driver_exe))
browser.get("http://172.16.15.18/prepay/login!init.do")
browser.implicitly_wait(100) #implicit wait
browser.maximize_window()
x1 = browser.find_element_by_id("czyId")
x1.send_keys("ChandpurAE1")
x2 = browser.find_element_by_id("pwd")
x2.send_keys("C6_029_Prepaid")
x3 = browser.find_element_by_xpath("//input[@type='button']")
x3.click()

for x in range(4):

    browser.find_elements_by_xpath('/html/body/div[1]/div/div/div[1]/div/div/div/div[3]/div/table/tbody/tr/td[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/em/button')[0].click()
    browser.find_elements_by_xpath('/html/body/div[6]/ul/li[1]/a/span')[0].click()
    sleep(2)
    browser.switch_to.frame(browser.find_element_by_id('mainFrame2'))
    meterNo = ws1.cell(row = 1+x, column = 1).value
    print("Meter No: ", meterNo)
    browser.switch_to.frame(browser.find_element_by_id('accountQueryIframe'))
    browser.find_element(By.ID, "metNo").send_keys(meterNo)
    browser.find_elements_by_class_name('ext_btn')[0].click()
    browser.switch_to_default_content()
    sleep(2) # necessary to make an iframe visible

    # frames = browser.find_elements_by_tag_name('iframe')

    # for frame in frames:
    #     print(frame)

    browser.switch_to.frame(browser.find_element_by_id('mainFrame2'))
    browser.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr/td[7]/button').click()
    sleep(2)
    browser.find_element_by_class_name('reissue').click()
    sleep(1)
    # frames = browser.find_elements_by_tag_name('iframe')

    # for frame in frames:
    #     print(frame)

    browser.switch_to_default_content()
    browser.switch_to.frame(browser.find_element_by_id('openwin'))
    # print("Hello")
    serial = browser.find_element_by_xpath('/html/body/table/tbody/tr[1]/td/table/tbody/tr[14]').text
    print("Token: ", serial)
    sequence = browser.find_element_by_xpath('/html/body/table/tbody/tr[1]/td/table/tbody/tr[11]').text
    print("Sequence: ", sequence[10:len(sequence)])
    ws1.cell(row = 1+x, column = 3).value = sequence[10:len(sequence)]
    ws1.cell(row = 1+x, column = 4).value = serial
    wb.save(os.path.join(os.getcwd(),excel_file))
    print("Ends: ", x+1)
    browser.refresh()


browser.close()