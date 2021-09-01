try:
    
    import os
    import shutil
    from time import sleep
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from openpyxl import load_workbook
    import logging
    import sys


    def consumer_fetch(excel_file):
        "Fetches Consumer ID from the Excel file provided"
        consumer_id=[]
        if excel_file.endswith('.xlsx')==False:
            excel_file=excel_file+".xlsx"
        if os.path.exists(os.path.join(os.getcwd(),excel_file))==False:
            print("EXCEL file not located in the same path as the program. Quitting ...")
            logging("EXCEL file not located in the same path as the program. Quitting ...")
            sleep(5)
            sys.exit()
        wb=load_workbook(filename=os.path.join(os.getcwd(),excel_file),read_only=True)
        ws=wb['Sheet 1']
        max_consumers=ws.max_row
        print("Total Customers count is {}".format(max_consumers-1))
        logging.info("Total Customers count is {}".format(max_consumers-1))
        for i in range(2,max_consumers+1):
            consumer_id.append(ws.cell(row=i,column=7).value)
        return consumer_id


    def bill_information(c_id):
        browser.switch_to.window(browser.window_handles[0])
        browser.get('http://180.211.137.22:8991/Pages/User/BillInformation.aspx')
        consumer_em=browser.find_element_by_id('cphMain_txtConsumer')
        consumer_em.send_keys(c_id)
        loc_code=browser.find_element_by_id('cphMain_txtLocationCode')
        loc_code.send_keys('b1')
        loc_code.send_keys(Keys.ENTER)
        sleep(5)
        browser.execute_script('window.scroll(0,100)')
        td=browser.find_elements_by_tag_name('td')
        last_month=td[8].text
        logging.info("Latest Month found from billInformation for {} is {}".format(c_id,last_month))
        print("Latest Month found from billInformation for {} is {}".format(c_id,last_month))
        return last_month


    def bill_print(c_id,last_month,report_count):
        browser.switch_to.window(browser.window_handles[1])
        browser.get('http://180.211.137.22:8991/Pages/User/BillPrint.aspx')
        consumer_em=browser.find_element_by_id('cphMain_txtConsumer')
        consumer_em.send_keys(c_id)
        loc_code=browser.find_element_by_id('cphMain_tbxLocation')
        loc_code.send_keys('b1')
        bill_cycle=browser.find_element_by_id('cphMain_txtBillCycle')
        bill_cycle.send_keys(last_month)
        bill_cycle.send_keys(Keys.TAB)
        gen_report_button=browser.find_element_by_id('cphMain_btnReport')
        gen_report_button.click()
        while len(os.listdir(os.path.join(os.getcwd(),'Reports')))==report_count:
            sleep(0.1)
        sleep(1)
        logging.info("Report prepared for consumer ID {} and the month {}".format(c_id,last_month))
        print("Report prepared for consumer ID {} and the month {}".format(c_id,last_month))


    def report_rename(c_id,last_month):
        files=os.listdir(os.path.join(os.getcwd(),'Reports'))
        files_full=[os.path.join(os.getcwd(),'Reports',i) for i in files]
        latest_file=max(files_full,key=os.path.getctime)
        new_name="Report"+"_"+str(c_id)+"_"+last_month+".pdf"
        os.rename(latest_file,os.path.join(os.getcwd(),'Reports',new_name))
        logging.info("Report saved for consumer ID {} and the month {}".format(c_id,last_month))
        print("Report saved for consumer ID {} and the month {}".format(c_id,last_month))




    log_format="%(asctime)s - %(levelname)s - %(message)s"
    logging.basicConfig(filename=os.path.join(os.getcwd(),"output.log"),filemode='a',format=log_format,level=logging.INFO)
    logging.info("Starting Operation...")


    chrome_options = webdriver.ChromeOptions()
    prefs = {'download.default_directory' : os.path.join(os.getcwd(),'Reports')}
    chrome_options.add_experimental_option('prefs', prefs)
    browser=webdriver.Chrome(executable_path=os.path.join(os.getcwd(),'chromedriver.exe'),options=chrome_options)
    browser.maximize_window()
    browser.get('http://180.211.137.22:8991/Pages/User/BillInformation.aspx')
    logging.info("Webpage BillInformation loaded.")
    browser.execute_script("window.open('http://180.211.137.22:8991/Pages/User/BillPrint.aspx')")
    logging.info("Webpage BillPrint loaded.")
    print("Both webpages loaded.")

        

    excel_file=input("Please input the EXCEL filename:")
    logging.info("EXCEL input file is %s",excel_file)
    consumer_id=consumer_fetch(excel_file)
    if len(consumer_id)>1:
        logging.info("CONSUMER ID array returned from consumer_fetch function.")

    if os.path.exists(os.path.join(os.getcwd(),'Reports')):
        shutil.rmtree(os.path.join(os.getcwd(),'Reports'))
        logging.info("Previous Reports folder deleted.")
        os.makedirs(os.path.join(os.getcwd(),'Reports'))
        logging.info("New Reports folder created.")



    for i in consumer_id:
        logging.info("Starting operation for {}".format(i))
        print("Starting operation for {}".format(i))
        last_month=bill_information(i)
        report_count=len(os.listdir(os.path.join(os.getcwd(),'Reports')))
        bill_print(i,last_month,report_count)
        report_rename(i,last_month)

    browser.execute_script("alert('Tasks Complete')")
    logging.info("Tasks Complete")
    print("Tasks Complete")

except Exception as e:
    logging.error("Exception Occured.",exc_info=True)
    print(e)
   
    
    
