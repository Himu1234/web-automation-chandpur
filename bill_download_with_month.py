import os
import re
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

dict = {
            "Jan" : 1,
            "Feb" : 2,
            "Mar" : 3,
            "Apr" : 4,
            "May" : 5,
            "Jun" : 6,
            "Jul" : 7,
            "Aug" : 8,
            "Sep" : 9,
            "Oct" : 10,
            "Nov" : 11,
            "Dec" : 12
}

excel_file = 'bill_print.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[1]]
max_consumers = ws1.max_row - 1
# max_consumers = 10

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))
count = 0
current_year = 21 # Last two digits of the current year. 
indent = 0 # Must check before each run
print(max_consumers-indent)

for x in range(max_consumers - indent - 1):
    
    clicks = current_year-int(ws1.cell(row = indent+2+x, column = 2).value[4:6]) 
    print("No of Clicks Required : ", clicks)
    print("Month : ", ws1.cell(row = indent+2+x, column = 2).value[0:3])
    xpath = '/html/body/div/div[2]/table/tbody/tr/td/span[{}]'.format(dict[ws1.cell(row = indent+2+x, column = 2).value[0:3]])
    print(xpath)
    
    browser.get('http://119.40.95.162:8991/Pages/User/BillPrint.aspx')
    browser.maximize_window()
    browser.implicitly_wait(100) #implicit wait
    x1 = browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = indent+2+x, column = 1).value
    print(cnum)
    x1.send_keys(cnum)
    x2 = browser.find_element_by_id("cphMain_tbxLocation")
    x2.send_keys("C6")
    x3 = browser.find_element_by_id("cphMain_txtBillCycle")
    x3.click()
    
    x6 = browser.find_element_by_xpath('/html/body/div/div[2]/table/thead/tr/th[1]')
    for i in range(clicks):
            x6.click()
            browser.implicitly_wait(100)
    
    x4 = browser.find_element_by_xpath(xpath)
    x4.click()
    x5 = browser.find_element_by_id('cphMain_btnReport')
    x5.click()
    
    browser.implicitly_wait(100) #implicit wait
    sleep(5)

browser.close()