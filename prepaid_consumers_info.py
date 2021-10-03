import os
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'ht_consumers_info.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(), excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[1]]
max_consumers = ws1.max_row
print(max_consumers)

browser = webdriver.Chrome(executable_path = os.path.join(os.getcwd(), driver_exe))

for x in range(max_consumers-1):

    browser.get("http://119.40.95.162:8991/Pages/User/ConsumerInfo.aspx")
    browser.implicitly_wait(100) #implicit wait
    browser.maximize_window()

    x1 = browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 2+x, column = 1).value
    x1.send_keys(cnum)
    # x2 = browser.find_element_by_id("cphMain_txtLocationCode")
    # x2.send_keys("C6")
    x3 = browser.find_element_by_id("cphMain_btnReport")
    x3.click()
    browser.implicitly_wait(100) #implicit wait
    sleep(2)
    td = browser.find_elements_by_tag_name('td')
    ws1.cell(row = 2+x, column = 2).value = td[1].text
    ws1.cell(row = 2+x, column = 3).value = td[5].text
    ws1.cell(row = 2+x, column = 4).value = td[13].text
    ws1.cell(row = 2+x, column = 5).value = td[15].text
    ws1.cell(row = 2+x, column = 6).value = td[17].text
    ws1.cell(row = 2+x, column = 7).value = td[19].text
    ws1.cell(row = 2+x, column = 8).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[2]/td[1]').text
    ws1.cell(row = 2+x, column = 9).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[2]').text
    ws1.cell(row = 2+x, column = 10).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[3]').text
    ws1.cell(row = 2+x, column = 11).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[13]').text
    ws1.cell(row = 2+x, column = 12).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[12]').text
    ws1.cell(row = 2+x, column = 13).value = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[3]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[11]').text
    print(cnum, "    ", td[1].text, "Address:    ", td[5].text, " Bill Group:   ", td[13].text, "Book No:    ", td[15].text)
    wb.save(os.path.join(os.getcwd(),excel_file))
    
wb.save(os.path.join(os.getcwd(),excel_file))
browser.close()