import os
import re
import math
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'Consumer_AC_No.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]
max_consumers = ws1.max_row - 2
print(max_consumers)

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))
clicks_for_starting_year = 19

for x in range(max_consumers):

    browser.get("http://180.211.137.22:8991/Login.aspx")
    browser.implicitly_wait(100) #implicit wait

    browser.find_element_by_id("cphMain_txxUserName").send_keys('xenb1')
    browser.find_element_by_id("cphMain_txxPassword").send_keys('12345')
    browser.find_element_by_id("cphMain_btnSubmit").click()
    browser.implicitly_wait(100)

    browser.find_element_by_xpath('/html/body/form/div[3]/div/div[1]/div/div[4]/div/ul/li[3]/a/i').click()
    browser.find_element_by_xpath('/html/body/form/div[3]/div/div[1]/div/div[4]/div/ul/li[3]/ul/li[2]/a').click()
    browser.implicitly_wait(100)

    con_no = ws1.cell(row = 10+x, column = 4).value
    browser.find_element_by_id("cphMain_txtConsumer").send_keys(con_no)
    browser.find_element_by_id("cphMain_tbxLocation").send_keys('B1')
    browser.find_element_by_id("cphMain_txtFromBillCycle").click()

    for i in range(clicks_for_starting_year):
        browser.find_element_by_xpath('/html/body/div[3]/div[2]/table/thead/tr/th[1]').click()

    browser.find_element_by_xpath('/html/body/div[3]/div[2]/table/tbody/tr/td/span[1]').click()
    browser.find_element_by_id("cphMain_txtToBillCycle").click()
    browser.find_element_by_xpath('/html/body/div[3]/div[2]/table/tbody/tr/td/span[12]').click()
    browser.implicitly_wait(100)
    print("Pre ....")
    browser.find_element_by_id("cphMain_btnReport").click()
    print("Post 1 ....")
    browser.implicitly_wait(100)
    print("Post 2 ....")
    sleep(5)
    print("Post 3 Sleep ended ....")
    print(con_no)
    browser.switch_to.window(browser.window_handles[0])    # In case of "no data found"

sleep(5)
wb.save(os.path.join(os.getcwd(),excel_file))
browser.close()