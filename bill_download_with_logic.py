import os
import re
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'B1_Dist_Feni_prepaid_Arr_upto_201912.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]
max_consumers = ws1.max_row - 1
print(max_consumers)

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))
count = 0

for x in range(max_consumers - 1):
    
    # print(ws1.cell(row = 2+x, column = 16).value, '    ', count)
    if ws1.cell(row = 2+x, column = 16).value >= 94.00 and ws1.cell(row = 2+x, column = 16).value <= 500.00 :
        
        browser.get("http://180.211.137.22:8991/Pages/User/BillInformation.aspx")
        browser.implicitly_wait(100) #implicit wait
        x1 = browser.find_element_by_id("cphMain_txtConsumer")
        cnum = ws1.cell(row = 2+x, column = 5).value
        # cnum = ws1.cell(row = 2+x, column = 2).value
        print(cnum, '    ', 'Arrear = ', ws1.cell(row = 2+x, column = 16).value, end = '    ')
        x1.send_keys(cnum)
        x2 = browser.find_element_by_id("cphMain_txtLocationCode")
        x2.send_keys("B1")
        x3 = browser.find_element_by_id('cphMain_btnReport')
        x3.click()
        x4 = browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[1]')
        month = x4.text
        print(month)       
        browser.get("http://180.211.137.22:8991/Pages/User/BillPrint.aspx")
        browser.implicitly_wait(100) #implicit wait
        x1 = browser.find_element_by_id("cphMain_txtConsumer")
        # print(cnum, '    ', 'Arrear = ', ws1.cell(row = 2+x, column = 15).value)
        x1.send_keys(cnum)
        x2 = browser.find_element_by_id("cphMain_tbxLocation")
        x2.send_keys("B1")
        x3 = browser.find_element_by_id("cphMain_txtBillCycle")
        x3.send_keys(month)   
        x4 = browser.find_element_by_xpath('/html/body/form/div[4]/div/div/div/div/div/div/div/div/div[3]/label')
        x4.click()
        x5 = browser.find_element_by_id('cphMain_btnReport')
        x5.click()
        browser.implicitly_wait(100) #implicit wait
        sleep(5)
        browser.switch_to.window(browser.window_handles[0])
        # count += 1
        
    else:
        print(ws1.cell(row = 2+x, column = 16).value)
        
wb.save(os.path.join(os.getcwd(),excel_file))
browser.quit()
