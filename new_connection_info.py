import os
import re
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'nc_tracking.xlsx'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[3]]
max_consumers = ws1.max_row
print(max_consumers)

driver_exe = 'chromedriver.exe'
browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

browser.get("http://119.40.95.163/Admin/Login")
browser.implicitly_wait(100)
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-content/md-input-container[1]/input').send_keys('ae1c6')
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-content/md-input-container[2]/input').send_keys('010203')
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-actions/button[1]').click()
browser.implicitly_wait(100)
browser.maximize_window()

for x in range(max_consumers):
    browser.refresh()
    sleep(2)
    trackingserial = ws1.cell(row = 1+x, column = 1).value
    browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[1]/md-input-container/input').send_keys(trackingserial)
    browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[1]/button').click()
    sleep(2)
    ws1.cell(row = 1+x, column = 2).value = browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[2]/table/tbody/tr[1]/th[1]').text
    ws1.cell(row = 1+x, column = 3).value = browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[2]/table/tbody/tr[2]/th[1]').text
    print(browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[2]/table/tbody/tr[1]/th[1]').text)
    print(browser.find_element_by_xpath('/html/body/section[2]/div/div/div[2]/div[2]/table/tbody/tr[2]/th[1]').text)
    wb.save(os.path.join(os.getcwd(),excel_file))


browser.close()
