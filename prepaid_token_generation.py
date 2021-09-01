import os
import re
import selenium
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


excel_file = 'B1_All_Prepaid_ARR_Consumer_LIst_201908.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

listMeter = ['10311108455', '10311108452', '10311108453', '10311065836']

browser.get("http://192.168.205.18/prepay/login!init.do;jsessionid=npFGRGZkearvpMUgCvkg.tomcat7_a")
browser.implicitly_wait(100) #implicit wait
browser.maximize_window()
x1=browser.find_element_by_id("czyId")
x1.send_keys("ChandpurAE1")
x2=browser.find_element_by_id("pwd")
x2.send_keys("C6_029_Prepaid")
x3=browser.find_element_by_xpath("//input[@type='button']")
x3.click()
print('Hello')
sleep(5)

for x in listMeter:

    browser.implicitly_wait(100)
    browser.get('http://192.168.205.18/prepay/prepay/mgtCode/codeMgt!ctc.do?timestamp=NaN&menuid=63100&menupath=Clear%20Tamper%20Status&curTabId=63100')  
    browser.implicitly_wait(100)
    generateBtn = browser.find_elements_by_class_name('ext_btn')[0]
    browser.switch_to.frame(browser.find_element_by_id('accountQueryIframe'))
    browser.implicitly_wait(100)
    browser.find_element(By.ID, "metNo").send_keys(x)
    print('1')
    browser.find_elements_by_class_name('ext_btn')[0].click()
    browser.implicitly_wait(100)
    print('2')
    browser.switch_to_default_content()
    sleep(2)
    generateBtn.click()
    browser.implicitly_wait(100)
    browser.find_element_by_xpath('/html/body/div[7]/div[2]/div[2]/div/div/div/div[1]/table/tbody/tr/td[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/em/button').click()
    # browser.find_element_by_xpath('/html/body/div[10]/div[2]/div[2]/div/div/div/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td/table').click()
    sleep(2)

browser.close()