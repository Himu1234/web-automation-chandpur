import os
import re
from typing import overload
import selenium
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

excel_file = 'Overdue_Check.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[1]]
max_consumers = ws1.max_row
print(max_consumers)

browser = webdriver.Chrome(executable_path = os.path.join(os.getcwd(), driver_exe))

browser.get("http://172.16.15.18/prepay/login!init.do")

browser.implicitly_wait(100) #implicit wait
browser.maximize_window()
x1 = browser.find_element_by_id("czyId")
x1.send_keys("ChandpurAE1")
x2 = browser.find_element_by_id("pwd")
x2.send_keys("C6_029_Prepaid")
x3 = browser.find_element_by_xpath("//input[@type='button']")
x3.click()
print('Hello')
sleep(5)

for x in range(max_consumers):

    browser.implicitly_wait(100)
    browser.get('http://172.16.15.18/prepay/prepay/mgtCode/codeMgt!ctc.do?timestamp=NaN&menuid=63100&menupath=Clear%20Tamper%20Status&curTabId=63100')  
    browser.implicitly_wait(100)
    browser.switch_to.frame(browser.find_element_by_id('accountQueryIframe'))
    browser.implicitly_wait(100)
    meterNo = ws1.cell(row = 1+x, column = 1).value
    print("Meter No: ", meterNo)
    browser.find_element(By.ID, "metNo").send_keys(meterNo)
    browser.find_elements_by_class_name('ext_btn')[0].click()
    browser.implicitly_wait(100)
    overdue = browser.find_element_by_id('dues').text
    print("Overdue: ", overdue)
    ws1.cell(row = 1+x, column = 3).value = overdue
    wb.save(os.path.join(os.getcwd(),excel_file))
    print('Ends : ', x+1)

browser.close()