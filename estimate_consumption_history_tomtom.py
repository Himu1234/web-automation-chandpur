import os
import re
import math
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'Check_Reading.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[2]]
max_consumers=ws1.max_row - 1
print(max_consumers)

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

for x in range(max_consumers):

    browser.get("http://180.211.137.22:8991/Pages/User/BillInformation.aspx")
    browser.implicitly_wait(100) #implicit wait

    x1=browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 2+x, column = 1).value
    if(cnum == None):
        print('None')
    else:
        x1.send_keys(cnum)
        x2=browser.find_element_by_id("cphMain_txtLocationCode")
        x2.send_keys("B1")
        x3=browser.find_element_by_id("cphMain_btnReport")
        x3.click()
        
        browser.implicitly_wait(100) #implicit wait
        
        x4=browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/label/select')
        x4.click()
        x5=browser.find_element_by_xpath('/html/body/form/div[4]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/label/select/option[4]')
        x5.click()
        
        browser.implicitly_wait(100) #implicit wait
        
        td=browser.find_elements_by_tag_name('td') 
        max_months = math.floor(len(td)/13)
        print(cnum, '    ', end = '')

        if(max_months > 12):   
            for i in range(12):
                if(x == 0):
                    ws1.cell(row = 1, column = 10+i).value = td[8+i*13].text
                ws1.cell(row = 2+x, column = 10+i).value = td[13+i*13].text
                print(td[13+i*13].text, '    ', end = '')
            ws1.cell(row = 2+x, column = 5).value = td[5].text
            ws1.cell(row = 2+x, column = 8).value = td[12].text
            ws1.cell(row = 2+x, column = 9).value = td[25].text
            ws1.cell(row = 2+x, column = 22).value = td[11].text
                
        else:
            for i in range(max_months):
                if(x == 0):
                    ws1.cell(row = 1, column = 10+i).value = td[8+i*13].text
                ws1.cell(row = 2+x, column = 10+i).value = td[13+i*13].text
                print(td[13+i*13].text, '    ', end = '')
            ws1.cell(row = 2+x, column = 5).value = td[5].text
            ws1.cell(row = 2+x, column = 8).value = td[12].text
            ws1.cell(row = 2+x, column = 9).value = td[25].text
            ws1.cell(row = 2+x, column = 22).value = td[11].text
        
        print()
        
wb.save(os.path.join(os.getcwd(),excel_file))
browser.quit()
