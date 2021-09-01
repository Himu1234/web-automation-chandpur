import os
import re
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

#excel_file = 'Prepaid_consumer_list.xlsx'
excel_file = 'Tomtoms_list.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]
#ws2 = wb[sheet[2]]

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

for x in range(33):

    browser.get("http://180.211.137.22:8991/Pages/User/BillInformation.aspx")
    browser.implicitly_wait(100) #implicit wait
    #browser.find_element_by_id("cphMain_txtConsumer")

    x1=browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 1+x, column = 1).value
    x1.send_keys(cnum)
    x2=browser.find_element_by_id("cphMain_txtLocationCode")
    x2.send_keys("B1")
    x3=browser.find_element_by_id("cphMain_btnReport")
    x3.click()
    
    #sleep(20) #Hard wait
    
    browser.implicitly_wait(100) #implicit wait
    
    # #wait = WebDriverWait(browser, 100)
    # #wait.until(ec.text_to_be_present_in_element((By.TAG_NAME, 'td'), 'Dec-18'))
    
    td=browser.find_elements_by_tag_name('td')
    month = td[8].text
    cons = td[13].text
    atbp = td[16].text
    mcond = td[11].text
    cname = td[5].text
    tariff = td[7].text
    # ws2.cell(row = 2+x, column = 1).value = td[14].text
    # ws2.cell(row = 2+x, column = 2).value = td[15].text
    # ws2.cell(row = 2+x, column = 3).value = td[16].text
    
    # browser.get("http://180.211.137.22:8991/Pages/User/BillPrint.aspx")
    # x4=browser.find_element_by_id("cphMain_txtConsumer")
    # x4.send_keys(ws1.cell(row = 1+x, column = 1).value)
    # x5=browser.find_element_by_id("cphMain_tbxLocation")
    # x5.send_keys("B1")
    # x6=browser.find_element_by_id("cphMain_txtBillCycle")
    #print(cnum, '   ', month, '   ',  cons, '      ', atbp, '      ', mcond)
    print(td[7].text)
    # x6.send_keys(month)
    # x6.send_keys(Keys.TAB)
    # x7=browser.find_element_by_id("cphMain_btnReport")
    # x7.click()
    # sleep(5)
    
# wb.save(os.path.join(os.getcwd(),excel_file))
sleep(10)
browser.close()