import os
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'chandpur_prepaid_arrear_202012.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[1]]
max_consumers = ws1.max_row
print(max_consumers)

browser = webdriver.Chrome(executable_path = os.path.join(os.getcwd(), driver_exe))

for x in range(max_consumers-1):

    browser.get("http://119.40.95.162:8991/Pages/User/BillInformation.aspx")
    browser.implicitly_wait(100) #implicit wait

    x1=browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 2+x, column = 1).value
    x1.send_keys(cnum)
    x2 = browser.find_element_by_id("cphMain_txtLocationCode")
    x2.send_keys("C6")
    x3 = browser.find_element_by_id("cphMain_btnReport")
    x3.click()
    browser.implicitly_wait(100) #implicit wait
    td = browser.find_elements_by_tag_name('td')
    ws1.cell(row = 2+x, column = 2).value = td[34].text
    ws1.cell(row = 2+x, column = 3).value = td[35].text
    ws1.cell(row = 2+x, column = 4).value = td[36].text
    ws1.cell(row = 2+x, column = 5).value = td[9].text
    print("Consumer No: ", cnum, "Last Payment Date: ", td[34].text, "Paid Amount: ", td[35].text, "Payment Date: ", td[36].text)
    
wb.save(os.path.join(os.getcwd(),excel_file))
browser.close()