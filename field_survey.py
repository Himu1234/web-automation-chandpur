import os
import re
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'saiful.xlsx'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]
max_consumers = ws1.max_row
print(max_consumers)

driver_exe = 'chromedriver.exe'
browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

browser.get("http://119.40.95.163/Admin/Login")
browser.implicitly_wait(100)
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-content/md-input-container[1]/input').send_keys('ae1c6')
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-content/md-input-container[2]/input').send_keys('c6_new_connection')
browser.find_element_by_xpath('/html/body/div/login/form/md-card/md-card-actions/button[1]').click()
browser.implicitly_wait(100)
browser.maximize_window()

for x in range(max_consumers):

    browser.get('http://119.40.95.163/FieldSurveyApprove/Index')
    browser.implicitly_wait(100)
    # print('hlw')
    # browser.find_element_by_xpath('/html/body/section[1]/aside[1]/div[2]/div/ul/li[3]/a').click()
    # browser.implicitly_wait(100)
    x1 = browser.find_element_by_xpath('/html/body/section[2]/div/div/trackingserial/div/div/md-input-container[1]/md-autocomplete/md-autocomplete-wrap/input')
    cnum = ws1.cell(row = 1+x, column = 1).value
    print("Tracking No: ", cnum)
    x1.send_keys(cnum)
    sleep(2)
    browser.find_element_by_xpath('/html/body/section[2]/div/div/trackingserial/div/div').click()
    sleep(2)
    browser.find_element_by_xpath('/html/body/section[2]/div/div/trackingserial/div/div/md-input-container[1]/md-autocomplete/md-autocomplete-wrap/input').click()
    sleep(2)
    browser.find_element_by_xpath('/html/body/md-virtual-repeat-container/div/div[2]/ul/li/md-autocomplete-parent-scope/span/span').click()
    sleep(2)
    browser.find_element_by_xpath('/html/body/section[2]/div/div/trackingserial/div/div/md-input-container[2]/md-select').click()
    browser.find_element_by_xpath('/html/body/div[6]/md-select-menu/md-content/md-option').click()
    browser.find_element_by_xpath('/html/body/section[2]/div/div/div/div/md-radio-group/md-radio-button[1]/div[1]/div[1]').click()
    browser.find_element_by_xpath('/html/body/section[2]/div/div/div/button').click()
    sleep(2)
    browser.find_element_by_xpath('/html/body/div[6]/div[7]/div/button').click()
    sleep(2)
    print('End: ', x+1)

browser.close()