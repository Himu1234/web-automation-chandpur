import os
import re
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'bill_print.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[3]]
max_consumers = ws1.max_row - 1
print(max_consumers)

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

for x in range(max_consumers):

    browser.get("http://119.40.95.162:8991/Pages/User/BillPrint.aspx")
    browser.implicitly_wait(100) #implicit wait
    browser.maximize_window()

    x1 = browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 1+x, column = 1).value
    x1.send_keys(cnum)
    x2 = browser.find_element_by_id("cphMain_tbxLocation")
    x2.send_keys("C6")
    x3 = browser.find_element_by_id("cphMain_txtBillCycle")
    x3.click()   
    x4 = browser.find_element_by_xpath('/html/body/div/div[2]/table/tbody/tr/td/span[3]')
    x4.click()
    x5 = browser.find_element_by_id('cphMain_btnReport')
    x5.click()
    
    browser.implicitly_wait(100) #implicit wait
    sleep(5)
        
wb.save(os.path.join(os.getcwd(),excel_file))
browser.quit()
