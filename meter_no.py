import os
import re
import math
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait

excel_file = 'bpdb.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename=os.path.join(os.getcwd(),excel_file), read_only=False)
sheet = wb.sheetnames
ws1 = wb[sheet[0]]
max_consumers=ws1.max_row - 1
print(max_consumers)

browser = webdriver.Chrome(executable_path=os.path.join(os.getcwd(), driver_exe))

for x in range(19):

    browser.get("http://180.211.137.22:8991/Pages/User/ConsumerInfo.aspx")
    browser.implicitly_wait(100) #implicit wait

    x1=browser.find_element_by_id("cphMain_txtConsumer")
    cnum = ws1.cell(row = 2+x, column = 1).value
    x1.send_keys(cnum)
    print(cnum)
    x3=browser.find_element_by_id("cphMain_btnReport")
    x3.click()
    
    browser.implicitly_wait(100) #implicit wait
    td=browser.find_elements_by_tag_name('td') 
    ws1.cell(row = 2+x, column = 2).value = td[28].text
    print(td[28].text)
    
wb.save(os.path.join(os.getcwd(),excel_file))
browser.close()