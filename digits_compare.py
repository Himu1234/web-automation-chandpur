import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook 
import pandas as pd

excel_file_data = 'digit_recognition.xlsx'
excel_file_submit = 'digit_recognition_submit.xlsx'

# convert csv to xlsx
# wb_data = Workbook()
# ws_data = wb_data.active
# with open('digit_recognition.csv', 'r') as f:
    # for row in csv.reader(f):
        # ws_data.append(row)
# wb_data.save(excel_file_data)

# wb_submit = Workbook()
# ws_submit = wb_submit.active
# with open('digit_recognition_submit.csv', 'r') as f:
    # for row in csv.reader(f):
        # ws_submit.append(row)
# wb_submit.save(excel_file_submit)



wb_data = load_workbook(filename=os.path.join(os.getcwd(),excel_file_data), read_only = False)
wb_submit = load_workbook(filename=os.path.join(os.getcwd(),excel_file_submit), read_only = False)
sheet_data = wb_data.sheetnames
sheet_submit = wb_submit.sheetnames
ws_data = wb_data[sheet_data[0]]
ws_submit = wb_submit[sheet_submit[0]]

for i in range(ws_data.max_row - 1):
    
    if(ws_data.cell(row = 2 + i, column = ws_data.max_column).value != ws_submit.cell(row = 2 + i, column = 2).value):
        print('Change detected!!')
        print('Previous value = ' + ws_submit.cell(row = 2 + i, column = 2).value)
        ws_submit.cell(row = 2 + i, column = 2).value = ws_data.cell(row = 2 + i, column = ws_data.max_column).value
        print('Current value = ' + ws_data.cell(row = 2 + i, column = ws_data.max_column).value)


# convert xlsx to csv
df = pd.read_excel("./digit_recognition_submit.xlsx")
df.to_csv("./digit_recognition_submit.csv", sep=",")