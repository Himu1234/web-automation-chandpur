import os
import re
import selenium
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.ui import WebDriverWait # Required for explicit wait
from selenium.webdriver.support import expected_conditions as ec # Required for explicit wait
from selenium.webdriver.common.by import By # Required for explicit wait
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

excel_file = 'token_generation_automation.xlsx'
driver_exe = 'chromedriver.exe'
wb = load_workbook(filename = os.path.join(os.getcwd(),excel_file), read_only = False)
sheet = wb.sheetnames
ws1 = wb[sheet[2]]
max_consumers = ws1.max_row

########################################################
########################################################
indent = 0 #Last valid iteration; Must check before each run
########################################################
########################################################

print(max_consumers-indent)

browser = webdriver.Chrome(executable_path = os.path.join(os.getcwd(), driver_exe))

browser.get("http://172.16.15.18/prepay/login!init.do")

browser.implicitly_wait(100) #implicit wait
browser.maximize_window()
x1 = browser.find_element_by_id("czyId")
x1.send_keys("ChandpurAE1")
x2 = browser.find_element_by_id("pwd")
x2.send_keys("C6_029_Prepaid")
x3 = browser.find_element_by_xpath("//input[@type='button']")
x3.click()
print('Hello')
sleep(5)

for x in range(max_consumers-indent):

    browser.implicitly_wait(100)
    browser.get('http://172.16.15.18/prepay/prepay/mgtCode/codeMgt!ctc.do?timestamp=NaN&menuid=63100&menupath=Clear%20Tamper%20Status&curTabId=63100')  
    browser.implicitly_wait(100)
    generateBtn = browser.find_elements_by_class_name('ext_btn')[0]
    selectBtn = browser.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/table/tbody/tr[2]/td[2]/select')
    selectOptn = browser.find_element_by_xpath('/html/body/table/tbody/tr/td[2]/form/table/tbody/tr[2]/td[2]/select/option[2]')
    browser.switch_to.frame(browser.find_element_by_id('accountQueryIframe'))
    browser.implicitly_wait(100)
    meterNo = ws1.cell(row = indent+1+x, column = 1).value
    print("Meter No: ", meterNo)
    browser.find_element(By.ID, "metNo").send_keys(meterNo)
    # print('1')
    browser.find_elements_by_class_name('ext_btn')[0].click()
    browser.implicitly_wait(100)
    # print('2')
    browser.switch_to_default_content()
    sleep(2)
    selectOptn.click()
    browser.implicitly_wait(100)
    selectBtn.click()
    browser.implicitly_wait(100)
    generateBtn.click()
    browser.implicitly_wait(100)
    browser.find_element_by_xpath('/html/body/div[7]/div[2]/div[2]/div/div/div/div[1]/table/tbody/tr/td[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/em/button').click()
    sleep(2)


    browser.switch_to.frame(browser.find_element_by_id('openwin'))
    serial = browser.find_element_by_xpath('/html/body/table/tbody/tr[1]/td/table/tbody/tr[14]').text
    print("Token: ", serial)
    sequence = browser.find_element_by_xpath('/html/body/table/tbody/tr[1]/td/table/tbody/tr[11]').text
    print("Sequence: ", sequence[10:len(sequence)])
    ws1.cell(row = indent+1+x, column = 3).value = sequence[10:len(sequence)]
    ws1.cell(row = indent+1+x, column = 4).value = serial


    ws1.cell(row = indent+1+x, column = 5).value = 'Done'
    wb.save(os.path.join(os.getcwd(),excel_file))
    print('Ends : ', x+1)

browser.close()